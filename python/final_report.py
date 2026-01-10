# final_report.py
# 2週間運用終了後の統計レポート生成

from datetime import datetime, timedelta
from sqlalchemy.orm import Session
from sqlalchemy import func, desc
from database import SessionLocal
from models import User, FoodLossRecord, LossReason
from statistics import get_week_boundaries
import json
from typing import Dict, List, Any


class FinalReportGenerator:
    """2週間運用終了後の最終レポート生成クラス"""
    
    def __init__(self):
        self.db = SessionLocal()
    
    def __del__(self):
        if hasattr(self, 'db'):
            self.db.close()
    
    def generate_complete_report(self) -> Dict[str, Any]:
        """完全な統計レポートを生成"""
        return {
            "report_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "user_statistics": self.get_user_statistics(),
            "reason_analysis": self.get_reason_analysis(), 
            "timeline_analysis": self.get_timeline_analysis(),
            "overall_summary": self.get_overall_summary(),
            "weekly_comparison": self.get_weekly_comparison(),
            "top_performers": self.get_top_performers(),
            "improvement_analysis": self.get_improvement_analysis()
        }
    
    def get_user_statistics(self) -> List[Dict[str, Any]]:
        """ユーザー別統計データ"""
        users = self.db.query(User).all()
        user_stats = []
        
        for user in users:
            # 総廃棄量
            total_weight = self.db.query(func.sum(FoodLossRecord.weight_grams))\
                .filter(FoodLossRecord.user_id == user.id).scalar() or 0
            
            # 廃棄回数
            record_count = self.db.query(func.count(FoodLossRecord.id))\
                .filter(FoodLossRecord.user_id == user.id).scalar() or 0
            
            # 平均廃棄量
            avg_weight = total_weight / record_count if record_count > 0 else 0
            
            # 最初と最後の記録日
            first_record = self.db.query(FoodLossRecord.record_date)\
                .filter(FoodLossRecord.user_id == user.id)\
                .order_by(FoodLossRecord.record_date.asc()).first()
            
            last_record = self.db.query(FoodLossRecord.record_date)\
                .filter(FoodLossRecord.user_id == user.id)\
                .order_by(FoodLossRecord.record_date.desc()).first()
            
            user_stats.append({
                "username": user.username,
                "email": user.email,
                "total_weight_grams": round(total_weight, 2),
                "record_count": record_count,
                "average_weight_grams": round(avg_weight, 2),
                "total_points": user.total_points,
                "first_record_date": first_record[0].strftime("%Y-%m-%d") if first_record else None,
                "last_record_date": last_record[0].strftime("%Y-%m-%d") if last_record else None,
                "participation_days": self._get_participation_days(user.id)
            })
        
        # 総廃棄量順にソート
        return sorted(user_stats, key=lambda x: x["total_weight_grams"], reverse=True)
    
    def get_reason_analysis(self) -> Dict[str, Any]:
        """廃棄理由別分析"""
        reason_stats = self.db.query(
            LossReason.reason_text,
            func.sum(FoodLossRecord.weight_grams).label('total_weight'),
            func.count(FoodLossRecord.id).label('count'),
            func.avg(FoodLossRecord.weight_grams).label('avg_weight')
        ).join(FoodLossRecord, LossReason.id == FoodLossRecord.loss_reason_id)\
         .group_by(LossReason.reason_text)\
         .order_by(desc('total_weight')).all()
        
        reasons = []
        total_all = sum(stat.total_weight for stat in reason_stats)
        
        for stat in reason_stats:
            percentage = (stat.total_weight / total_all * 100) if total_all > 0 else 0
            reasons.append({
                "reason": stat.reason_text,
                "total_weight_grams": round(stat.total_weight, 2),
                "count": stat.count,
                "average_weight_grams": round(stat.avg_weight, 2),
                "percentage": round(percentage, 1)
            })
        
        return {
            "reason_breakdown": reasons,
            "most_common_reason": reasons[0]["reason"] if reasons else None,
            "total_reasons": len(reasons)
        }
    
    def get_timeline_analysis(self) -> Dict[str, Any]:
        """時系列分析（日別・週別）"""
        # 日別統計
        daily_stats = self.db.query(
            func.date(FoodLossRecord.record_date).label('date'),
            func.sum(FoodLossRecord.weight_grams).label('total_weight'),
            func.count(FoodLossRecord.id).label('count')
        ).group_by(func.date(FoodLossRecord.record_date))\
         .order_by('date').all()
        
        daily_data = []
        for stat in daily_stats:
            daily_data.append({
                "date": stat.date.strftime("%Y-%m-%d"),
                "total_weight_grams": round(stat.total_weight, 2),
                "record_count": stat.count
            })
        
        return {
            "daily_statistics": daily_data,
            "total_days_with_records": len(daily_data),
            "average_daily_waste": round(sum(d["total_weight_grams"] for d in daily_data) / len(daily_data), 2) if daily_data else 0
        }
    
    def get_overall_summary(self) -> Dict[str, Any]:
        """全体サマリー"""
        # 全体統計
        total_weight = self.db.query(func.sum(FoodLossRecord.weight_grams)).scalar() or 0
        total_records = self.db.query(func.count(FoodLossRecord.id)).scalar() or 0
        total_users = self.db.query(func.count(User.id)).scalar() or 0
        total_points = self.db.query(func.sum(User.total_points)).scalar() or 0
        
        # 参加率計算
        active_users = self.db.query(func.count(func.distinct(FoodLossRecord.user_id))).scalar() or 0
        participation_rate = (active_users / total_users * 100) if total_users > 0 else 0
        
        return {
            "total_waste_grams": round(total_weight, 2),
            "total_records": total_records,
            "total_users": total_users,
            "active_users": active_users,
            "participation_rate_percent": round(participation_rate, 1),
            "total_points_awarded": total_points,
            "average_waste_per_user": round(total_weight / active_users, 2) if active_users > 0 else 0,
            "average_records_per_user": round(total_records / active_users, 1) if active_users > 0 else 0
        }
    
    def get_weekly_comparison(self) -> Dict[str, Any]:
        """週別比較（1週目 vs 2週目）"""
        today = datetime.now()
        
        # 現在の週
        current_week_start, current_week_end = get_week_boundaries(today)
        
        # 1週間前の週
        last_week = today - timedelta(weeks=1)
        last_week_start, last_week_end = get_week_boundaries(last_week)
        
        # 各週のデータ取得
        def get_week_data(start_date, end_date):
            records = self.db.query(FoodLossRecord)\
                .filter(FoodLossRecord.record_date >= start_date)\
                .filter(FoodLossRecord.record_date <= end_date).all()
            
            total_weight = sum(r.weight_grams for r in records)
            unique_users = len(set(r.user_id for r in records))
            
            return {
                "total_weight_grams": round(total_weight, 2),
                "record_count": len(records),
                "active_users": unique_users,
                "average_per_user": round(total_weight / unique_users, 2) if unique_users > 0 else 0
            }
        
        week1_data = get_week_data(last_week_start, last_week_end)
        week2_data = get_week_data(current_week_start, current_week_end)
        
        # 改善率計算
        improvement_rate = 0
        if week1_data["total_weight_grams"] > 0:
            improvement_rate = ((week1_data["total_weight_grams"] - week2_data["total_weight_grams"]) 
                              / week1_data["total_weight_grams"] * 100)
        
        return {
            "week1": {
                "period": f"{last_week_start.strftime('%Y-%m-%d')} ~ {last_week_end.strftime('%Y-%m-%d')}",
                **week1_data
            },
            "week2": {
                "period": f"{current_week_start.strftime('%Y-%m-%d')} ~ {current_week_end.strftime('%Y-%m-%d')}",
                **week2_data
            },
            "improvement_rate_percent": round(improvement_rate, 1),
            "is_improving": improvement_rate > 0
        }
    
    def get_top_performers(self) -> Dict[str, Any]:
        """優秀者・改善者ランキング"""
        # ポイント獲得ランキング
        top_points = self.db.query(User.username, User.total_points)\
            .order_by(desc(User.total_points)).limit(5).all()
        
        # 廃棄量削減ランキング（週別比較で計算）
        # ここでは簡略化して総廃棄量が少ない順
        user_waste = self.db.query(
            User.username,
            func.sum(FoodLossRecord.weight_grams).label('total_waste')
        ).join(FoodLossRecord, User.id == FoodLossRecord.user_id)\
         .group_by(User.username)\
         .order_by('total_waste').limit(5).all()
        
        return {
            "top_points_earners": [
                {"username": user.username, "points": user.total_points}
                for user in top_points
            ],
            "least_waste_producers": [
                {"username": user.username, "total_waste_grams": round(user.total_waste, 2)}
                for user in user_waste
            ]
        }
    
    def get_improvement_analysis(self) -> Dict[str, Any]:
        """改善効果分析"""
        # 簡単な改善指標
        weekly_comparison = self.get_weekly_comparison()
        overall = self.get_overall_summary()
        
        # 予想される年間削減効果
        if weekly_comparison["week1"]["total_weight_grams"] > 0:
            weekly_reduction = (weekly_comparison["week1"]["total_weight_grams"] - 
                              weekly_comparison["week2"]["total_weight_grams"])
            annual_projection = weekly_reduction * 52  # 年間52週
        else:
            annual_projection = 0
        
        return {
            "weekly_reduction_grams": round(weekly_reduction, 2) if 'weekly_reduction' in locals() else 0,
            "projected_annual_reduction_grams": round(annual_projection, 2),
            "projected_annual_reduction_kg": round(annual_projection / 1000, 2),
            "behavior_change_indicator": "改善傾向" if weekly_comparison["is_improving"] else "要注意",
            "engagement_score": round(overall["participation_rate_percent"] * 
                                    (overall["average_records_per_user"] / 10), 1)
        }
    
    def _get_participation_days(self, user_id: int) -> int:
        """ユーザーの参加日数を計算"""
        days = self.db.query(func.count(func.distinct(func.date(FoodLossRecord.record_date))))\
            .filter(FoodLossRecord.user_id == user_id).scalar()
        return days or 0
    
    def export_to_excel(self, filename: str = None) -> str:
        """レポートをExcelファイルに出力"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            print("⚠️ Excelサポートには pandas と openpyxl が必要です:")
            print("pip install pandas openpyxl")
            return None
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"food_loss_report_{timestamp}.xlsx"
        
        report = self.generate_complete_report()
        
        # Excelワークブック作成
        wb = Workbook()
        
        # 1. 全体サマリーシート
        ws_summary = wb.active
        ws_summary.title = "全体サマリー"
        
        # ヘッダースタイル
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        
        summary_data = [
            ["項目", "値", "単位"],
            ["総廃棄量", report["overall_summary"]["total_waste_grams"], "g"],
            ["総廃棄量(kg)", round(report["overall_summary"]["total_waste_grams"]/1000, 2), "kg"],
            ["総記録数", report["overall_summary"]["total_records"], "件"],
            ["参加者数", report["overall_summary"]["active_users"], "人"],
            ["登録者数", report["overall_summary"]["total_users"], "人"],
            ["参加率", report["overall_summary"]["participation_rate_percent"], "%"],
            ["総獲得ポイント", report["overall_summary"]["total_points_awarded"], "P"],
            ["ユーザー平均廃棄量", report["overall_summary"]["average_waste_per_user"], "g"],
        ]
        
        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_num, column=col_num, value=value)
                if row_num == 1:  # ヘッダー行
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
        
        # 列幅調整
        for column in ws_summary.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws_summary.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # 2. ユーザー別統計シート
        ws_users = wb.create_sheet(title="ユーザー別統計")
        
        # データフレーム作成
        users_df = pd.DataFrame(report["user_statistics"])
        
        # ヘッダー名を日本語に変更
        column_mapping = {
            "username": "ユーザー名",
            "email": "メールアドレス", 
            "total_weight_grams": "総廃棄量(g)",
            "record_count": "記録回数",
            "average_weight_grams": "平均廃棄量(g)",
            "total_points": "獲得ポイント",
            "first_record_date": "初回記録日",
            "last_record_date": "最終記録日",
            "participation_days": "参加日数"
        }
        users_df = users_df.rename(columns=column_mapping)
        
        # Excelに書き込み
        for r in dataframe_to_rows(users_df, index=False, header=True):
            ws_users.append(r)
        
        # ヘッダースタイル適用
        for cell in ws_users[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # 列幅自動調整
        for column in ws_users.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws_users.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # 3. 廃棄理由分析シート
        ws_reasons = wb.create_sheet(title="廃棄理由分析")
        
        reasons_df = pd.DataFrame(report["reason_analysis"]["reason_breakdown"])
        reason_column_mapping = {
            "reason": "廃棄理由",
            "total_weight_grams": "総廃棄量(g)",
            "count": "回数",
            "average_weight_grams": "平均廃棄量(g)",
            "percentage": "割合(%)"
        }
        reasons_df = reasons_df.rename(columns=reason_column_mapping)
        
        for r in dataframe_to_rows(reasons_df, index=False, header=True):
            ws_reasons.append(r)
        
        for cell in ws_reasons[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        for column in ws_reasons.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws_reasons.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # 4. 週別比較シート
        ws_weekly = wb.create_sheet(title="週別比較")
        
        weekly = report["weekly_comparison"]
        weekly_data = [
            ["項目", "1週目", "2週目", "差分"],
            ["期間", weekly["week1"]["period"], weekly["week2"]["period"], ""],
            ["廃棄量(g)", weekly["week1"]["total_weight_grams"], weekly["week2"]["total_weight_grams"], 
             weekly["week1"]["total_weight_grams"] - weekly["week2"]["total_weight_grams"]],
            ["記録数", weekly["week1"]["record_count"], weekly["week2"]["record_count"],
             weekly["week1"]["record_count"] - weekly["week2"]["record_count"]],
            ["参加者数", weekly["week1"]["active_users"], weekly["week2"]["active_users"],
             weekly["week1"]["active_users"] - weekly["week2"]["active_users"]],
            ["改善率(%)", "", "", weekly["improvement_rate_percent"]],
            ["状況", "", "", "改善中" if weekly["is_improving"] else "要注意"]
        ]
        
        for row_num, row_data in enumerate(weekly_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = ws_weekly.cell(row=row_num, column=col_num, value=value)
                if row_num == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
        
        for column in ws_weekly.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws_weekly.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # 5. 日別統計シート
        ws_daily = wb.create_sheet(title="日別統計")
        
        daily_df = pd.DataFrame(report["timeline_analysis"]["daily_statistics"])
        daily_column_mapping = {
            "date": "日付",
            "total_weight_grams": "廃棄量(g)",
            "record_count": "記録数"
        }
        daily_df = daily_df.rename(columns=daily_column_mapping)
        
        for r in dataframe_to_rows(daily_df, index=False, header=True):
            ws_daily.append(r)
        
        for cell in ws_daily[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        for column in ws_daily.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws_daily.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # ファイル保存
        wb.save(filename)
        return filename
        """レポートをJSONファイルに出力"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"final_report_{timestamp}.json"
        
        report = self.generate_complete_report()
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        
        return filename
    
    def print_summary_report(self):
        """コンソールにサマリーレポートを出力"""
        report = self.generate_complete_report()
        
        print("=" * 60)
        print("🍽️  食品ロス削減プロジェクト - 2週間運用レポート")
        print("=" * 60)
        print(f"📅 レポート生成日時: {report['report_date']}")
        print()
        
        # 全体サマリー
        overall = report["overall_summary"]
        print("📊 【全体サマリー】")
        print(f"   総廃棄量: {overall['total_waste_grams']}g ({overall['total_waste_grams']/1000:.2f}kg)")
        print(f"   総記録数: {overall['total_records']}件")
        print(f"   参加者数: {overall['active_users']}/{overall['total_users']}人 ({overall['participation_rate_percent']}%)")
        print(f"   総獲得ポイント: {overall['total_points_awarded']}P")
        print()
        
        # 週別比較
        weekly = report["weekly_comparison"]
        print("📈 【週別比較】")
        print(f"   1週目: {weekly['week1']['total_weight_grams']}g")
        print(f"   2週目: {weekly['week2']['total_weight_grams']}g")
        print(f"   改善率: {weekly['improvement_rate_percent']}%")
        print(f"   状況: {'✅ 改善中' if weekly['is_improving'] else '⚠️ 要注意'}")
        print()
        
        # 廃棄理由TOP3
        reasons = report["reason_analysis"]["reason_breakdown"][:3]
        print("🗑️ 【主な廃棄理由 TOP3】")
        for i, reason in enumerate(reasons, 1):
            print(f"   {i}. {reason['reason']}: {reason['total_weight_grams']}g ({reason['percentage']}%)")
        print()
        
        # 優秀者
        top_points = report["top_performers"]["top_points_earners"][:3]
        print("🏆 【ポイント獲得 TOP3】")
        for i, user in enumerate(top_points, 1):
            print(f"   {i}. {user['username']}: {user['points']}P")
        print()
        
        # 改善効果予測
        improvement = report["improvement_analysis"]
        print("🔮 【改善効果予測】")
        print(f"   年間削減予測: {improvement['projected_annual_reduction_kg']}kg")
        print(f"   エンゲージメントスコア: {improvement['engagement_score']}/100")
        print("=" * 60)


def main():
    """メイン実行関数"""
    print("2週間運用統計レポートを生成中...")
    
    generator = FinalReportGenerator()
    
    # コンソール出力
    generator.print_summary_report()
    
    # Excelファイル出力
    excel_filename = generator.export_to_excel()
    if excel_filename:
        print(f"\n📊 Excelレポートを保存しました: {excel_filename}")
    
    # JSONファイル出力
    json_filename = generator.export_to_json()
    print(f"📁 詳細レポート(JSON)を保存しました: {json_filename}")
    
    print("\n✅ レポート生成完了！")
    print("\n💡 管理者向け:")
    print(f"   📈 分析用: {excel_filename}")
    print(f"   🔧 技術用: {json_filename}")


if __name__ == "__main__":
    main()